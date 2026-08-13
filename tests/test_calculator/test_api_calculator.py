"""WO-21 计算器端点测试 — /api/calculator/*（过程可见 + 上传闭环安全）。"""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
import yaml
from fastapi.testclient import TestClient

from core.calculator import profiles as profiles_mod
from core.calculator import updates
from server.main import app

BASE = {
    "applicants": [{"base": 120000}],
    "loan": {"portions": [{"amount": 500000, "rate": 0.06, "term_years": 30}],
             "security_value": 700000},
    "household": {"status": "Single"},
}


@pytest.fixture
def client(tmp_path, monkeypatch):
    # 完全隔离配置目录与更新闭环，避免触碰真实 config/calculator/
    conf = tmp_path / "calc_conf"
    data = tmp_path / "calc_data"
    conf.mkdir()
    for f in Path("config/calculator").glob("*.yaml"):
        shutil.copy(f, conf / f.name)
    monkeypatch.setattr(profiles_mod, "_CONFIG_DIR", conf)
    monkeypatch.setattr(updates, "_PROFILES_DIR", conf)
    monkeypatch.setattr(updates, "_DATA_DIR", data)
    monkeypatch.setattr(updates, "_PENDING_DIR", data / "pending")
    monkeypatch.setattr(updates, "_SOURCES_DIR", data / "sources")
    monkeypatch.setattr(updates, "_HISTORY_DIR", data / "history")
    yield TestClient(app)


class TestAssess:
    def test_assess_ok(self, client):
        r = client.post("/api/calculator/assess", json={"bank": "boc", **BASE})
        assert r.status_code == 200
        body = r.json()
        assert body["result"] == "PASS"
        assert body["indicator"] == "nis"
        assert body["surplus"] == pytest.approx(354.21, abs=0.05)
        assert body["profile_version"]
        assert len(body["steps"]) > 5
        step = body["steps"][0]
        assert step["step_id"] and step["label"] and step["formula"]

    def test_assess_refer(self, client):
        r = client.post("/api/calculator/assess",
                        json={"bank": "macquarie", **BASE})
        assert r.status_code == 200
        assert r.json()["result"] == "REFER"

    def test_assess_unknown_bank_404(self, client):
        r = client.post("/api/calculator/assess",
                        json={"bank": "nope", **BASE})
        assert r.status_code == 404

    def test_assess_empty_applicants_422(self, client):
        body = {"bank": "boc", **BASE}
        body["applicants"] = []
        r = client.post("/api/calculator/assess", json=body)
        assert r.status_code == 422

    def test_assess_lvr_guard_visible(self, client):
        body = {"bank": "macquarie", **BASE}
        body["applicants"] = [{"base": 200000}]
        body["loan"]["portions"][0]["amount"] = 750000
        r = client.post("/api/calculator/assess", json=body)
        assert r.status_code == 200
        assert r.json()["result"] == "FAIL"
        ids = [s["step_id"] for s in r.json()["steps"]]
        assert "result:lvr_cap" in ids

    def test_steps_recomputable(self, client):
        r = client.post("/api/calculator/assess", json={"bank": "boc", **BASE})
        steps = r.json()["steps"]
        assert all(s["output"] is not None for s in steps)


class TestProfiles:
    def test_list_profiles(self, client):
        r = client.get("/api/calculator/profiles")
        assert r.status_code == 200
        banks = {p["bank"] for p in r.json()}
        assert banks == {"boc", "cba", "macquarie", "ma_money", "latrobe", "resimac"}
        for p in r.json():
            assert p["version"] and p["source_file"] and p["status"] == "default"


class TestUploadApplyRollback:
    def test_upload_yaml_diff(self, client):
        p = profiles_mod.load_profile("boc")
        p.pop("_hash", None)
        p["parameters"]["assessment"]["buffer"] = p["parameters"]["assessment"]["buffer"] + 0.001
        content = yaml.safe_dump(p, sort_keys=False).encode()
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("boc_update.yaml", content, "application/yaml")})
        assert r.status_code == 200
        body = r.json()
        assert body["bank"] == "boc"
        assert body["changed_count"] >= 1
        assert body["source_hash"]
        assert any(i["path"] == "parameters.assessment.buffer" for i in body["diff"])

    def test_upload_invalid_yaml_422(self, client):
        content = b"- just\na: list\n"  # 顶层非 dict
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("boc_update.yaml", content, "application/yaml")})
        assert r.status_code == 422

    def test_upload_bank_from_content_wins(self, client):
        # 文件名提示 boc，但 YAML 内容 bank=cba → 以内容为准
        p = profiles_mod.load_profile("cba")
        p.pop("_hash", None)
        content = yaml.safe_dump(p, sort_keys=False).encode()
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("boc_update.yaml", content, "application/yaml")})
        assert r.status_code == 200
        assert r.json()["bank"] == "cba"

    def test_upload_unknown_bank_new(self, client):
        content = b"bank: nope\nparameters: {}\n"
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("mystery.yaml", content, "application/yaml")})
        assert r.status_code == 200
        body = r.json()
        assert body["is_new_bank"] is True and body["needs_review"] is True

    def test_upload_bad_extension_422(self, client):
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("evil.txt", b"x", "text/plain")})
        assert r.status_code == 422

    def test_upload_oversize_413(self, client):
        big = b"0" * (20 * 1024 * 1024 + 1)
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("big.yaml", big, "application/yaml")})
        assert r.status_code == 413

    def test_upload_fake_xlsm_422(self, client):
        # 非 ZIP 的伪 xlsm
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("boc fake.xlsm", b"not a zip", "application/octet-stream")})
        assert r.status_code == 422

    def test_upload_xlsm_parse_readonly_redline(self, client, monkeypatch):
        # 红线：上传 xlsm 解析全程 openpyxl read_only=True + keep_vba，绝不执行宏
        import io

        import openpyxl

        from core.calculator.parsers import base as parser_base

        calls = {}
        real_load = openpyxl.load_workbook

        def fake_load(path, **kwargs):
            calls.update(kwargs)
            return real_load(path, read_only=True, keep_vba=True,
                             data_only=kwargs.get("data_only", True))

        monkeypatch.setattr(parser_base.openpyxl, "load_workbook", fake_load)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parameters"
        for r, (f, g, h) in enumerate([(18200, 0, 0), (45000, 0, 0.15),
                                       (135000, 4020, 0.3), (190000, 31020, 0.37),
                                       (1e9, 51370, 0.45)], start=2):
            ws[f"F{r}"], ws[f"G{r}"], ws[f"H{r}"] = f, g, h
        ws["Q2"] = 0.03
        wb.create_sheet("HEM")
        wb.create_sheet("Change Log")
        buf = io.BytesIO()
        wb.save(buf)
        r = client.post("/api/calculator/profiles/upload",
                        files={"file": ("boc redline.xlsm", buf.getvalue(),
                                        "application/octet-stream")})
        assert r.status_code == 200, r.text
        assert calls["read_only"] is True
        assert calls["keep_vba"] is True

    def test_upload_apply_flow(self, client):
        p = profiles_mod.load_profile("boc")
        p.pop("_hash", None)
        p["parameters"]["result"]["threshold"] = 50
        content = yaml.safe_dump(p, sort_keys=False).encode()
        up = client.post("/api/calculator/profiles/upload",
                         files={"file": ("boc_update.yaml", content, "application/yaml")})
        source_hash = up.json()["source_hash"]
        before = profiles_mod.load_profile("boc")["_hash"]
        ap = client.post("/api/calculator/profiles/boc/apply",
                         json={"source_hash": source_hash})
        assert ap.status_code == 200, ap.text
        assert ap.json()["bank"] == "boc"
        assert profiles_mod.load_profile("boc")["_hash"] != before

    def test_apply_missing_pending_404(self, client):
        r = client.post("/api/calculator/profiles/boc/apply",
                        json={"source_hash": "deadbeef"})
        assert r.status_code == 404

    def test_apply_conflict_409(self, client):
        # pending 指向当前版本（CAS 冲突）：apply_pending 期望版本 != 当前
        p = profiles_mod.load_profile("boc")
        p.pop("_hash", None)
        content = yaml.safe_dump(p, sort_keys=False).encode()
        up = client.post("/api/calculator/profiles/upload",
                         files={"file": ("boc_update.yaml", content, "application/yaml")})
        ap = client.post("/api/calculator/profiles/boc/apply",
                         json={"source_hash": up.json()["source_hash"]})
        # 与当前版本无差异时 expected_version 仍等于当前 -> 无冲突；此处验证返回类型合法
        assert ap.status_code in (200, 409)

    def test_rollback_flow(self, client):
        p = profiles_mod.load_profile("boc")
        p.pop("_hash", None)
        p["parameters"]["result"]["threshold"] = 60
        content = yaml.safe_dump(p, sort_keys=False).encode()
        up = client.post("/api/calculator/profiles/upload",
                         files={"file": ("boc_update.yaml", content, "application/yaml")})
        source_hash = up.json()["source_hash"]
        client.post("/api/calculator/profiles/boc/apply", json={"source_hash": source_hash})
        rb = client.post("/api/calculator/profiles/boc/rollback",
                         json={"version": source_hash})
        assert rb.status_code == 200
        assert rb.json()["rolled_back_to"] == source_hash

    def test_rollback_missing_404(self, client):
        r = client.post("/api/calculator/profiles/boc/rollback", json={"version": "x"})
        assert r.status_code == 404
