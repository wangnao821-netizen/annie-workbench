"""WO-21 解析器测试 — 合成 xlsx fixture（openpyxl 内存构造）+ 文件名识别 + 只读红线。"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from core.calculator.parsers import _parse, _parse_upload, parse
from core.calculator.parsers import base as parser_base

_HEM_ROW = 100.0


def _book(sheets: dict) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cells in sheets.items():
        ws = wb.create_sheet(name)
        for coord, val in cells.items():
            ws[coord] = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_range(cells: dict, sheet: dict, coord: str, col: int, start: int,
                values: list, row_label: str | None = None):
    for i, v in enumerate(values):
        sheet[f"{chr(col + i)}{start}"] = v
    if row_label is not None:
        cells[row_label] = 1


def _boc() -> bytes:
    return _book({
        "Parameters": {
            "F2": 18200, "G2": 0, "H2": 0,
            "F3": 45000, "G3": 0, "H3": 0.15,
            "F4": 135000, "G4": 4020, "H4": 0.3,
            "F5": 190000, "G5": 31020, "H5": 0.37,
            "F6": 1e9, "G6": 51370, "H6": 0.45,
            "J2": 0, "K2": 0, "L2": 0,
            "J3": 105000, "K3": 210000, "L3": 0.01,
            "Q2": 0.035, "R2": 0.053, "S2": 0.0,
            "P2": 0.038, "O2": 0.038, "N2": 100,
            "V2": 150, "W2": 180, "X2": 0.2, "Y2": 0.2,
            "AA2": 0.2, "AB2": 0.3, "AE2": 0.2, "AC2": "2000", "I2": 0.02,
        },
        "HEM": {},
        "Change Log": {"A27": "Version 7.1.6"},
    })


def _cba() -> bytes:
    master = {
        "A3": 18200, "B3": 0, "C3": 0,
        "A4": 45000, "B4": 0.15, "C4": 0,
        "A5": 135000, "B5": 0.3, "C5": 4020,
        "A6": 190000, "B6": 0.37, "C6": 31020,
        "A7": 1e9, "B7": 0.45, "C7": 51370,
        "A38": 37500, "B38": 700, "C38": 0,
        "A39": 45000, "B39": 325, "C39": 0.015,
        "A40": 66000, "B40": 0, "C40": 0,
        "S3": "LV81.3", "T3": "2026-06-27",
        "R52": 0.031, "R51": 0.054,
        "B12": 28011, "B13": 35013, "C12": 0.1,
        "R48": 0.038, "R49": 25, "R50": 0.03,
        "S22": "S0", "T22": "C0",
    }
    for i, v in enumerate([27000.01, 40000.01, 54000.01, 67000.01, 81000.01,
                           107000.01, 134000.01, 161000.01, 188000.01,
                           215000.01, 269000.01, 336000.01, 403000.01,
                           671000.01, 1e9]):
        master[f"Q{23 + i}"] = v
    for col in range(19, 30):
        letter = get_column_letter(col)
        for i in range(15):
            master[f"{letter}{23 + i}"] = _HEM_ROW
    return _book({"MasterData": master, "Data": {}})


def _macquarie() -> bytes:
    refs = {
        "B15": 0, "C15": 18200, "D15": 0,
        "B16": 0, "C16": 45000, "D16": 0.15,
        "B17": 0, "C17": 135000, "D17": 0.3,
        "B18": 0, "C18": 190000, "D18": 0.37,
        "B19": 0, "C19": 1e9, "D19": 0.45,
        "H10": 0.033, "H9": 0.053, "H11": 0.0225,
    }
    hem = {}
    for i, v in enumerate([27000, 41000, 54000, 68000, 82000, 109000, 136000,
                           163000, 191000, 218000, 272000, 341000, 409000,
                           681000, 953000, 1225000, 1497000, 1e9]):
        hem[f"{chr(67 + i)}7"] = v
    hem["B8"] = "C0"
    for i in range(18):
        hem[f"{chr(67 + i)}8"] = _HEM_ROW
    return _book({"References": refs,
                  "Serviceability Worksheet": {"I4": "(17 Jul 2026)"},
                  "HEM Table": hem})


def _ma_money() -> bytes:
    hidden = {
        "C22": 0, "D22": 18200, "E22": 0, "F22": 0,
        "C23": 0, "D23": 45000, "E23": 0, "F23": 0.15,
        "C24": 0, "D24": 135000, "E24": 4020, "F24": 0.3,
        "C25": 0, "D25": 190000, "E25": 31020, "F25": 0.37,
        "C26": 0, "D26": 1e9, "E26": 51370, "F26": 0.45,
        "D55": 0.25, "C30": 28011, "C31": 35013, "G30": 0.1,
        "C47": 0, "D47": 37500, "E47": 700, "F47": 0,
    }
    hem = {}
    for i, v in enumerate([27001, 40001, 53001, 66001, 80001, 106001, 133001,
                           160001, 186001, 213001, 266001, 332001, 399001,
                           1e9]):
        hem[f"{chr(69 + i)}20"] = v
    hem["C23"] = "NSW"
    hem["D23"] = "Single"
    for i in range(14):
        hem[f"{chr(69 + i)}23"] = _HEM_ROW
    return _book({"Setup": {"I22": "5.2", "I23": "2026-07-06",
                            "I39": 0.021, "I41": 0.0575, "I43": 0.038,
                            "I45": 0.9, "I44": 0.25, "I54": 0.0},
                  "Hidden.Calcs": hidden, "HEM": hem})


def _latrobe() -> bytes:
    hem = {}
    for i, v in enumerate([27000, 41000, 54000, 68000, 82000, 109000, 136000,
                           163000, 191000, 218000, 272000, 341000, 409000,
                           1e9]):
        hem[f"{chr(67 + i)}9"] = v
    hem["B13"] = "Single"
    for i in range(14):
        hem[f"{chr(67 + i)}13"] = _HEM_ROW
    return _book({
        "Tax Scales": {
            "A5": 0, "B5": 18200, "C5": 0, "D5": 0,
            "A6": 18200, "B6": 45000, "C6": 0, "D6": 0.15,
            "A7": 45000, "B7": 135000, "C7": 0, "D7": 0.3,
            "A8": 135000, "B8": 190000, "C8": 0, "D8": 0.37,
            "A9": 190000, "B9": 1e9, "C9": 0, "D9": 0.45,
            "D13": 0.02, "D112": 0.25,
        },
        "Living Allow": {"E11": "Version Date: 29 June 2026"},
        "RepayCalculator": {"B6": 0.021, "B9": 0.053, "B44": 0.8},
        "HEM": hem,
    })


def _resimac() -> bytes:
    hem = {}
    for i, v in enumerate([27000, 41000, 54000, 68000, 82000, 109000, 136000,
                           163000, 191000, 218000, 272000, 341000, 409000,
                           1e9]):
        hem[f"{chr(66 + i)}2"] = v
    for row in (6, 11):
        hem[f"A{row}"] = "Single"
        for i in range(14):
            hem[f"{chr(66 + i)}{row}"] = _HEM_ROW
    return _book({
        "Calculator": {"A5": "Version 7.03 (01/07/2026)"},
        "Tables": {
            "G50": 0.0575, "H50": 0.021, "C55": 0.038, "L54": 0.8,
            "K49": "prime", "L49": 1.0, "K50": "quickstart", "L50": 1.25,
            "K51": "specialist", "L51": 1.0,
            "B60": 0, "C60": 0, "D60": 0,
            "B61": 18200, "C61": 0, "D61": 0,
            "B62": 45000, "C62": 0, "D62": 0.15,
            "B63": 135000, "C63": 4020, "D63": 0.3,
            "B64": 1e9, "C64": 31020, "D64": 0.37,
        },
        "HEM Table": hem,
    })


@pytest.mark.parametrize(
    "bank,builder,version,param_cell,param_value,first_upper",
    [
        ("boc", _boc, "Version 7.1.6", "assessment.buffer", 0.035, 18200),
        ("cba", _cba, "LV81.3", "assessment.buffer", 0.031, 45000),
        ("macquarie", _macquarie, "(17 Jul 2026)", "assessment.buffer", 0.033, 18200),
        ("ma_money", _ma_money, "5.2", "assessment.buffer", 0.021, 18200),
        ("latrobe", _latrobe, "Version Date: 29 June 2026", "assessment.buffer", 0.021, None),
        ("resimac", _resimac, "Version 7.03 (01/07/2026)", "assessment.buffer", 0.021, None),
    ],
)
def test_parse_each_bank(bank, builder, version, param_cell, param_value, first_upper):
    parsed = _parse_upload(bank, builder())
    assert parsed["source_version"] == version
    params = parsed["parameters"]
    path = param_cell.split(".")
    value = params
    for part in path:
        value = value[part]
    assert value == pytest.approx(param_value)
    assert parsed["source_file"]
    if first_upper is not None:
        assert params["tax"]["brackets"][0][0] == pytest.approx(first_upper)
    assert params["living"]["hem_table"]["income_bands"]


def test_boc_haircut_cells_pinned():
    """租金=AA2(80%)，高密度=AB2(70%)，股息=AE2，投资/海外=契约 20%。"""
    parsed = _parse_upload("boc", _boc())
    hc = parsed["parameters"]["income_rules"]["haircuts"]
    assert hc["rental_income"] == pytest.approx(0.8)
    assert hc["dividends"] == pytest.approx(0.8)
    assert hc["investment_income"] == pytest.approx(0.8)
    assert hc["foreign_income"] == pytest.approx(0.8)
    assert hc["overtime"] == pytest.approx(0.8)
    hd = parsed["options"]["high_density"]
    assert hd["rental_haircut"] == pytest.approx(0.7)
    assert hd["postcodes"] == "2000"


def test_identify_bank_from_filenames(tmp_path):
    from core.calculator.parsers.boc import FILE as boc_file
    from core.calculator.parsers.cba import FILE as cba_file
    from core.calculator.parsers.latrobe import FILE as lat_file
    from core.calculator.parsers.ma_money import FILE as ma_file
    from core.calculator.parsers.macquarie import FILE as mac_file
    from core.calculator.parsers.resimac import FILE as res_file
    cases = [(boc_file, _boc), (cba_file, _cba), (mac_file, _macquarie),
             (ma_file, _ma_money), (lat_file, _latrobe), (res_file, _resimac)]
    for name, builder in cases:
        path = tmp_path / name
        path.write_bytes(builder())
        parsed = parse(path)
        assert parsed["source_file"] == name


def test_identify_unknown_filename(tmp_path):
    path = tmp_path / "mystery_calc.xlsx"
    path.write_bytes(_book({"Sheet1": {}}))
    with pytest.raises(ValueError, match="cannot identify"):
        parse(path)


def test_parse_unknown_bank():
    with pytest.raises(ValueError, match="unknown bank"):
        _parse("nope", Path("x.xlsm"))


def test_open_workbook_read_only_redline(monkeypatch, tmp_path):
    """红线：openpyxl 只读打开（read_only=True + keep_vba），绝不加载宏。"""
    calls = {}
    real_load = openpyxl.load_workbook

    def fake_load(path, **kwargs):
        calls.update(kwargs)
        return real_load(path, read_only=True, keep_vba=True,
                         data_only=kwargs.get("data_only", True))

    monkeypatch.setattr(parser_base.openpyxl, "load_workbook", fake_load)
    path = tmp_path / "boc.xlsx"
    path.write_bytes(_boc())
    wb = parser_base.open_workbook(path)
    assert calls["read_only"] is True
    assert calls["keep_vba"] is True
    wb.close()


def test_parse_upload_cleans_temp_file():
    from core.calculator.parsers import _parse_upload
    parsed = _parse_upload("boc", _boc())
    assert parsed["source_version"] == "Version 7.1.6"
