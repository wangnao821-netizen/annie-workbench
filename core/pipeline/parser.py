"""File parser for Phase 1C using LiteParse.

Integrates the LiteParse package for high-speed, local document parsing.
Returns a unified ParseResult with parsed text and metadata.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import extract_msg
from liteparse import LiteParse

from core.logger import get_logger

logger = get_logger(__name__)


@dataclass
class ParseResult:
    """Result of parsing a single file."""

    text: str
    text_quality: str  # "high" | "medium" | "low"
    tables: list[dict[str, object]] = field(default_factory=list)
    metadata: dict[str, object] = field(default_factory=dict)
    attachments: list[str] = field(default_factory=list)
    parse_route: str = "native_text"  # "native_text" | "ocr_text" | "ocr_failed"
    business_fields: dict[str, str] = field(default_factory=dict)


class ParseError(Exception):
    """Raised when file parsing absolutely fails."""


def _worker_parse_liteparse(file_path_str: str) -> tuple[str, bool, int]:
    lp = LiteParse(output_format="markdown", ocr_enabled=True)
    res = lp.parse(Path(file_path_str))
    text = res.text or ""
    is_ocr = getattr(res, "ocr_used", None) is True or getattr(res, "is_scanned", None) is True
    num_pages = getattr(res, "num_pages", 0)
    return text, is_ocr, num_pages


def _parse_with_liteparse(file_path: Path) -> tuple[str, bool, int]:
    """Process-isolated LiteParse call to protect Python runtime from Rust/C FFI crashes."""
    from concurrent.futures import ProcessPoolExecutor
    with ProcessPoolExecutor(max_workers=1) as executor:
        future = executor.submit(_worker_parse_liteparse, str(file_path))
        return future.result(timeout=60)


def parse_file(file_path: Path) -> ParseResult:
    """Parse a file using LiteParse and return a ``ParseResult``.

    Args:
        file_path: Path to the file to parse (any supported type).

    Returns:
        A ``ParseResult`` mapping the LiteParse output.

    Raises:
        ParseError: If the file does not exist.
    """
    if not file_path.exists():
        raise ParseError(f"File not found: {file_path}")

    # Special handling for Outlook .msg files
    if file_path.suffix.lower() == ".msg":
        logger.info("Parsing .msg file using extract-msg for %s", file_path.name)
        try:
            msg = extract_msg.Message(str(file_path))

            # Extract basic email fields
            subject = msg.subject or "No Subject"
            sender = msg.sender or "Unknown Sender"
            date = msg.date or "Unknown Date"
            body = msg.body or ""

            # Format text cleanly
            text = f"Subject: {subject}\nFrom: {sender}\nDate: {date}\n\n{body}"

            # Get attachment names (we don't extract them physically here yet, just list them)
            attachments = [att.longFilename or att.shortFilename for att in msg.attachments]

            return ParseResult(
                text=text,
                text_quality="high",
                parse_route="native_text",
                metadata={"engine": "extract-msg", "subject": subject, "sender": sender, "date": str(date)},
                attachments=attachments
            )
        except Exception as e:
            logger.error("extract-msg failed on %s: %s", file_path, e)
            return ParseResult(
                text="",
                text_quality="low",
                parse_route="ocr_failed",
                metadata={"engine": "extract-msg", "error": str(e)},
            )

    # Direct text reading for .txt and .md files
    if file_path.suffix.lower() in (".txt", ".md"):
        try:
            content_str = file_path.read_text(encoding="utf-8", errors="ignore")
            return ParseResult(
                text=content_str,
                text_quality="high",
                parse_route="native_text",
                metadata={"engine": "native_text", "file_size": file_path.stat().st_size}
            )
        except Exception as e:
            logger.error("Failed to read text file %s: %s", file_path, e)

    # Native Python extraction for Word (.docx) & Excel (.xlsx / .xlsm)
    ext = file_path.suffix.lower()
    if ext in (".docx", ".doc"):
        try:
            import docx
            doc = docx.Document(str(file_path))
            paras = [p.text for p in doc.paragraphs if p.text.strip()]
            full_text = "\n".join(paras)
            return ParseResult(
                text=full_text,
                text_quality="high",
                parse_route="native_docx",
                metadata={"engine": "python-docx", "paragraphs": len(paras)}
            )
        except Exception as e:
            logger.warning("python-docx extraction failed for %s: %s", file_path.name, e)

    if ext in (".xlsx", ".xlsm", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            text_lines = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_str = " | ".join(str(c) for c in row if c is not None)
                    if row_str.strip():
                        text_lines.append(row_str)
            full_text = "\n".join(text_lines)
            return ParseResult(
                text=full_text,
                text_quality="high",
                parse_route="native_excel",
                metadata={"engine": "openpyxl", "sheets": len(wb.sheetnames)}
            )
        except Exception as e:
            logger.warning("openpyxl extraction failed for %s: %s", file_path.name, e)

    if ext in (".png", ".jpg", ".jpeg", ".bmp", ".webp"):
        # 优先尝试 LiteParse OCR 提取图片中的文字（贷款文档常有手机拍照）
        ocr_text = ""
        try:
            lp_text, is_ocr, _ = _parse_with_liteparse(file_path)
            # 有效字符判定：过滤 markdown 代码块包装（```text\n\n```）等空白噪音，
            # 要求至少 10 个字母/数字/中文字符才算有效 OCR 文本
            if lp_text and sum(1 for ch in lp_text if ch.isalnum()) >= 10:
                ocr_text = lp_text.strip()
        except Exception as ocr_err:
            logger.debug("Image OCR via LiteParse failed (non-critical): %s", ocr_err)

        if ocr_text:
            return ParseResult(
                text=ocr_text,
                text_quality="medium",
                parse_route="image_ocr",
                metadata={"format": ext, "ocr_chars": len(ocr_text)},
            )

        # Fallback: 图片 OCR 文字过少或失败，返回尺寸占位信息
        try:
            from PIL import Image
            img = Image.open(str(file_path))
            w, h = img.size
        except Exception:
            w, h = 0, 0
        return ParseResult(
            text=f"[图片文件 {file_path.name} | 尺寸: {w}x{h} px | OCR 未识别到有效文字]",
            text_quality="low",
            parse_route="image_metadata",
            metadata={"width": w, "height": h, "format": ext},
        )

    logger.info("Initializing LiteParse for %s", file_path.name)

    text = ""
    is_ocr = False
    num_pages = 0

    # Attempt 1: LiteParse with 1 retry on timeout/crash
    for attempt in range(2):
        try:
            text, is_ocr, num_pages = _parse_with_liteparse(file_path)
            break
        except Exception as e:
            if attempt == 0:
                logger.warning(
                    "LiteParse attempt 1 failed on %s (will retry): %s",
                    file_path.name, e,
                )
                continue
            logger.error("LiteParse failed on %s after 2 attempts: %s", file_path, e)

    # If LiteParse produced no text, try PDF-specific fallbacks
    if not text.strip() and ext == ".pdf":
        # Fallback A: pypdf (fast, good for native-text PDFs)
        try:
            import pypdf
            reader = pypdf.PdfReader(str(file_path))
            num_pages = len(reader.pages)
            pdf_text = "\n".join([page.extract_text() or "" for page in reader.pages])
            if pdf_text.strip():
                text = pdf_text
                is_ocr = False
                logger.info("pypdf fallback succeeded for %s (%d chars)", file_path.name, len(text))
        except Exception as pe:
            logger.error("pypdf fallback failed on %s: %s", file_path, pe)

    if not text.strip() and ext == ".pdf":
        # Fallback B: pdfplumber (better for tables/complex layouts)
        try:
            import pdfplumber
            with pdfplumber.open(str(file_path)) as pdf:
                num_pages = len(pdf.pages)
                pages_text = []
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        pages_text.append(page_text)
                if pages_text:
                    text = "\n".join(pages_text)
                    is_ocr = False
                    logger.info("pdfplumber fallback succeeded for %s (%d chars)", file_path.name, len(text))
        except ImportError:
            logger.debug("pdfplumber not installed, skipping fallback")
        except Exception as pb_err:
            logger.error("pdfplumber fallback failed on %s: %s", file_path, pb_err)

    # If still too short for a PDF, it's likely a scanned document — try image OCR
    if len(text.strip()) < 50 and ext == ".pdf":
        logger.info("PDF text too short (%d chars), attempting image OCR for %s", len(text.strip()), file_path.name)
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(file_path))
            ocr_texts = []
            for page_idx in range(min(len(doc), 10)):  # Cap at 10 pages for cost
                pix = doc[page_idx].get_pixmap(dpi=200)
                img_bytes = pix.tobytes("png")
                # Write temp image and OCR via LiteParse
                import tempfile
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp.write(img_bytes)
                    tmp_path = Path(tmp.name)
                try:
                    lp_text, _, _ = _parse_with_liteparse(tmp_path)
                    if lp_text and len(lp_text.strip()) > 10:
                        ocr_texts.append(lp_text.strip())
                except Exception:
                    pass
                finally:
                    tmp_path.unlink(missing_ok=True)
            doc.close()

            if ocr_texts:
                text = "\n\n".join(ocr_texts)
                is_ocr = True
                logger.info("Image OCR fallback succeeded for %s (%d chars from %d pages)",
                            file_path.name, len(text), len(ocr_texts))
        except ImportError:
            logger.debug("PyMuPDF (fitz) not installed, skipping image OCR fallback")
        except Exception as ocr_err:
            logger.warning("Image OCR fallback failed for %s: %s", file_path.name, ocr_err)

    # Final: no text at all → report failure
    if not text.strip():
        return ParseResult(
            text="",
            text_quality="low",
            parse_route="ocr_failed",
            metadata={"engine": "all_fallbacks_exhausted", "pages": num_pages},
        )

    # Quality assessment
    if len(text.strip()) < 50:
        quality = "low"
        route = "ocr_failed" if is_ocr else "native_text"
    else:
        quality = "high" if not is_ocr else "medium"
        route = "ocr_text" if is_ocr else "native_text"

    return ParseResult(
        text=text,
        text_quality=quality,
        parse_route=route,
        metadata={"engine": "liteparse", "pages": num_pages, "ocr_used": is_ocr},
    )
