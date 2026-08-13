"""WO-21 解析器基类：read_only 安全打开 + 版本锚点 + HEM 表抽取。

红线（契约 §四）：openpyxl 仅 read_only=True 读取缓存值，keep_vba 保留宏
但绝不执行；解析永不写源文件。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl


def open_workbook(path: Path, data_only: bool = True):
    """契约红线：read_only=True（不加载/执行 VBA 宏）、keep_vba=True。"""
    return openpyxl.load_workbook(path, read_only=True, keep_vba=True,
                                  data_only=data_only)


def cell(ws, coordinate: str) -> Any:
    try:
        return ws[coordinate].value
    except (KeyError, TypeError, AttributeError, ValueError):
        return None


def family_code(label: Any, prefix: str | None = None) -> str:
    """'Couple with 1 child'/'Single, 2 dep'/'S3' -> C1/S2/S3。"""
    text = str(label or "").strip().lower()
    if re.match(r"^[cs]\d+$", text):
        return text.upper()
    if prefix is None:
        prefix = "C" if "couple" in text or "2 adults" in text else "S"
    m = re.search(r"(\d+)", text)
    if m:
        return f"{prefix}{int(m.group(1))}"
    return f"{prefix}0"


def numeric_row(ws, row: int, min_col: int, max_col: int) -> list[float]:
    values: list[float] = []
    for col in range(min_col, max_col + 1):
        v = ws.cell(row=row, column=col).value
        values.append(float(v) if isinstance(v, (int, float)) else None)
    return values


def label_and_codes(labels: list[Any]) -> list[tuple[str, str]]:
    """把标签列转为 (code, 原文标签)。'add adult/dep' 映射 S_add/C_add/C_adult。"""
    out: list[tuple[str, str]] = []
    for label in labels:
        text = str(label or "").strip()
        low = text.lower()
        prefix = "C" if "couple" in low else "S"
        if "add" in low and ("adult" in low):
            out.append((f"{prefix}_adult", text))
        elif "add" in low:
            out.append((f"{prefix}_add", text))
        else:
            out.append((family_code(label), text))
    return out
